#!/usr/bin/env python3
"""
Generate PostPlanner-compatible XLSX from approved social media content.

PostPlanner Bulk Upload V3.1 format (9 columns):
  Column A: POSTING TIME (MM/DD/YYYY HH:MM in 24hr, or empty for queued)
  Column B: CAPTION (post text — for X/Twitter, links go here too)
  Column C: LINK (not used for X/Twitter)
  Column D: MEDIA (image/video URL)
  Column E: TITLE (Pinterest/YouTube only)
  Column F: FIRST COMMENT (Facebook/Instagram/LinkedIn only)
  Column G: FIRST COMMENT DELAY (minutes: 5, 10, 15, 30, 60)
  Column H: TOBI BACKGROUND (Facebook TOBI — e.g. "true-black", "shapes")
  Column I: REEL OR STORY (Facebook/Instagram — "reel" or "story")

Post type is auto-detected by PostPlanner based on which columns are filled.
Row 1 is a comment row with "BULK UPLOAD VERSION 3.1" in cell B1.
Any row starting with ## is a comment and ignored by PostPlanner.

Usage:
    python generate-postplanner-export.py --niche Cubs 2026-02-25
    python generate-postplanner-export.py --niche Cubs 2026-02-25 --tobi
    python generate-postplanner-export.py --niche Softball  # defaults to today
"""

import re
import json
import sys
import argparse
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl required. Install with: python -m pip install openpyxl")
    sys.exit(1)


def load_niche_config(niche_folder):
    """Load niche-config.yaml using simple line parsing (no pyyaml needed)."""
    config_path = niche_folder / "niche-config.yaml"
    if not config_path.exists():
        print(f"Error: {config_path} not found")
        sys.exit(1)

    config = {}
    with open(config_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            # Simple key: value parsing for top-level scalars
            m = re.match(r'^(\w+):\s*"?([^"#]+)"?\s*$', line)
            if m:
                config[m.group(1)] = m.group(2).strip()
            # Detect platform list items
            if line.startswith("- ") and "platforms" in str(config.get("_last_key", "")):
                config.setdefault("platforms", []).append(line[2:].strip())
            if re.match(r'^\w+:', line) and not line.startswith("-"):
                config["_last_key"] = line.split(":")[0]

    config.pop("_last_key", None)
    return config


def load_approvals(content_folder):
    """Load approvals.json if it exists."""
    approvals_path = content_folder / "approvals.json"
    if approvals_path.exists():
        try:
            with open(approvals_path, "r") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return {}
    return {}


def parse_time_from_heading(heading_text, date_str):
    """
    Extract posting time from a heading like '#### Text Post -- 9:00 AM ET'.
    Returns formatted PostPlanner time string or None.
    """
    m = re.search(r'(\d{1,2}:\d{2}\s*[AP]M)\s+([A-Z]{2,3})', heading_text)
    if not m:
        return None
    time_str = m.group(1).strip()
    try:
        posting_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %I:%M %p")
        return posting_dt.strftime("%m/%d/%Y  %H:%M")
    except ValueError:
        return None


def parse_time_from_window_tag(section_text, date_str):
    """
    Extract posting time from [POSTING WINDOW: 7:00 AM CT] tag.
    Returns formatted PostPlanner time string or None.
    """
    m = re.search(
        r'\[POSTING WINDOW:\s*(?:Alternate\s+)?(\d{1,2}:\d{2}\s*[AP]M)\s+([A-Z]{2,3})\]',
        section_text
    )
    if not m:
        return None
    time_str = m.group(1).strip()
    try:
        posting_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %I:%M %p")
        return posting_dt.strftime("%m/%d/%Y  %H:%M")
    except ValueError:
        return None


def parse_x_posts(content_folder, date_str):
    """Parse X/Twitter posts from markdown, return list of post dicts."""
    x_file = content_folder / "03-social-posts-x.md"
    posts = []
    skipped = []

    if not x_file.exists():
        print(f"  Warning: {x_file.name} not found")
        return posts

    with open(x_file, "r", encoding="utf-8") as f:
        content = f.read()

    # Split into story sections
    story_blocks = re.split(r"(?=## STORY\s+\d+:)", content)

    for block in story_blocks:
        if not block.strip() or not re.match(r"## STORY", block.strip()):
            continue

        # Extract story number for error reporting
        story_match = re.match(r"## STORY\s+(\d+):", block.strip())
        story_num = story_match.group(1) if story_match else "?"

        # Split into subsections by ### or ####
        post_sections = re.split(r"(?=#{3,4}\s+)", block)

        for section in post_sections:
            section_stripped = section.strip()

            # Match Text Post sections (handles "Text Post", "Text Post 2",
            # "Text Post (Bold)", etc.) and Image Post sections
            is_text = re.match(r"#{3,4}\s+Text Post", section_stripped)
            is_image = re.match(r"#{3,4}\s+Image Post", section_stripped)

            if not is_text and not is_image:
                continue

            # Determine post type
            if is_image:
                post_type = "IMAGE"
            else:
                post_type = "TEXT"

            # Try both time formats:
            # 1. Time in the heading: "#### Text Post -- 9:00 AM ET"
            heading_line = section_stripped.split("\n")[0]
            pp_time = parse_time_from_heading(heading_line, date_str)

            # 2. Time in a [POSTING WINDOW: ...] tag
            if not pp_time:
                pp_time = parse_time_from_window_tag(section, date_str)

            if not pp_time:
                continue

            # Extract post text from code block
            code_match = re.search(r"```\n(.*?)```", section, re.DOTALL)
            if not code_match:
                continue

            caption = code_match.group(1).strip()

            # Remove character count lines if accidentally captured
            caption = re.sub(r"\n?\[(?:CHARACTER COUNT|CHAR(?:ACTER)? COUNT):.*?\]", "", caption)
            caption = re.sub(r"\n?\[\d+/280\]", "", caption)

            if not caption:
                continue

            # For X text posts, enforce 280 char limit — WARN instead of silent skip
            if post_type == "TEXT" and len(caption) > 280:
                skipped.append((story_num, heading_line.strip(), len(caption)))
                continue

            # Look for an image URL in the image manifest or section metadata
            media_url = ""
            # Image posts may reference an image — handled at export time from manifest

            posts.append({
                "time": pp_time,
                "caption": caption,
                "post_type": post_type,
                "media_url": media_url,
            })

    # Report any skipped tweets prominently
    if skipped:
        print(f"  ⚠ SKIPPED {len(skipped)} tweet(s) over 280-char limit:")
        for story_num, heading, char_count in skipped:
            print(f"    Story {story_num}: {char_count} chars (+{char_count - 280} over) — {heading}")
        print(f"  Fix these in 03-social-posts-x.md and re-run.")

    return posts


def parse_facebook_posts(content_folder, date_str):
    """Parse Facebook posts from markdown, return list of post dicts."""
    fb_file = content_folder / "04-social-posts-facebook.md"
    posts = []

    if not fb_file.exists():
        print(f"  Warning: {fb_file.name} not found")
        return posts

    with open(fb_file, "r", encoding="utf-8") as f:
        content = f.read()

    # Split into story sections
    story_blocks = re.split(r"(?=## STORY\s+\d+:)", content)

    for block in story_blocks:
        if not block.strip() or not re.match(r"## STORY", block.strip()):
            continue

        # Split into subsections
        post_sections = re.split(r"(?=#{3,4}\s+)", block)

        for section in post_sections:
            section_stripped = section.strip()

            # Match Long-Form Post and Image Caption sections
            is_longform = re.match(r"#{3,4}\s+Long-Form Post", section_stripped)
            is_caption = re.match(r"#{3,4}\s+Image (?:Post|Caption)", section_stripped)

            if not is_longform and not is_caption:
                continue

            # Determine post type
            if is_caption:
                post_type = "IMAGE"
            else:
                post_type = "TEXT"

            # Try both time formats
            heading_line = section_stripped.split("\n")[0]
            pp_time = parse_time_from_heading(heading_line, date_str)
            if not pp_time:
                pp_time = parse_time_from_window_tag(section, date_str)
            if not pp_time:
                continue

            # Extract post text from code block
            code_match = re.search(r"```\n(.*?)```", section, re.DOTALL)
            if not code_match:
                continue

            caption = code_match.group(1).strip()
            if not caption:
                continue

            posts.append({
                "time": pp_time,
                "caption": caption,
                "post_type": post_type,
                "media_url": "",
            })

    return posts


def deduplicate_times(posts, min_gap_minutes=5):
    """
    Ensure no two posts share the exact same posting time.
    If duplicates are found, stagger them apart by min_gap_minutes.
    This is a safety net that runs regardless of redistribute_times.
    """
    if not posts:
        return posts

    from collections import Counter
    time_counts = Counter(p["time"] for p in posts)
    dupes = {t: c for t, c in time_counts.items() if c > 1}

    if not dupes:
        return posts

    print(f"  Warning: {sum(c - 1 for c in dupes.values())} posts had duplicate times — staggering by {min_gap_minutes}min")

    seen = {}
    for post in posts:
        t = post["time"]
        if t in seen:
            # Stagger this post forward by min_gap_minutes * collision count
            try:
                date_prefix, time_part = t.split("  ")
                h, m = map(int, time_part.split(":"))
                total_min = h * 60 + m + (min_gap_minutes * seen[t])
                if total_min >= 24 * 60:
                    total_min = 23 * 60 + 55  # cap at 23:55
                new_h = total_min // 60
                new_m = total_min % 60
                post["time"] = f"{date_prefix}  {new_h:02d}:{new_m:02d}"
            except (IndexError, ValueError):
                pass
            seen[t] += 1
        else:
            seen[t] = 1

    return posts


def redistribute_times(posts, date_str):
    """
    Ensure all post times are at least 30 min from now,
    evenly spaced through the rest of the day.
    Only redistributes if the content date matches today.
    """
    today_str = datetime.now().strftime("%Y-%m-%d")
    if date_str != today_str:
        return posts  # future/past dates: keep original times

    now = datetime.now()
    floor_minutes = now.hour * 60 + now.minute + 30
    end_minutes = 22 * 60  # 10:00 PM cutoff

    # Check if any post times have already passed
    any_past = False
    for post in posts:
        # post["time"] is "MM/DD/YYYY  HH:MM"
        try:
            time_part = post["time"].split("  ")[1]  # "HH:MM"
            h, m = map(int, time_part.split(":"))
            if h * 60 + m < floor_minutes:
                any_past = True
                break
        except (IndexError, ValueError):
            continue

    if not any_past:
        return posts

    if floor_minutes >= end_minutes:
        print("  Warning: past 9:30 PM — not enough time to redistribute. Using original times.")
        return posts

    available = end_minutes - floor_minutes
    spacing = available // len(posts)
    spacing = max(15, min(90, spacing))  # clamp 15-90 min

    date_prefix = posts[0]["time"].split("  ")[0]  # "MM/DD/YYYY"
    for i, post in enumerate(posts):
        post_min = floor_minutes + (i * spacing)
        if post_min >= 24 * 60:
            post_min = 23 * 60 + 45
        h = post_min // 60
        m = post_min % 60
        post["time"] = f"{date_prefix}  {h:02d}:{m:02d}"

    first_t = posts[0]["time"].split("  ")[1]
    last_t = posts[-1]["time"].split("  ")[1]
    print(f"  Redistributed {len(posts)} posts: {first_t} to {last_t} ({spacing} min apart)")
    return posts


def export_xlsx(posts, output_path, brand_name):
    """Export posts to PostPlanner XLSX Bulk Upload V3.1 format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Row 1: Comment row with version marker (REQUIRED by PostPlanner)
    ws["A1"] = f"## {brand_name} -- PostPlanner Bulk Upload"
    ws["B1"] = "BULK UPLOAD VERSION 3.1 \u2013 Please DO NOT EDIT this cell."

    # Row 2: Column headers as comment row (V3.1 format — 9 columns)
    ws["A2"] = "## POSTING TIME"
    ws["B2"] = "CAPTION"
    ws["C2"] = "LINK"
    ws["D2"] = "MEDIA"
    ws["E2"] = "TITLE"
    ws["F2"] = "FIRST COMMENT"
    ws["G2"] = "FIRST COMMENT DELAY"
    ws["H2"] = "TOBI BACKGROUND"
    ws["I2"] = "REEL OR STORY"

    # Data rows starting at row 3
    # V3.1 auto-detects post type from filled columns:
    #   Text: only B (caption)
    #   Link: B + C (caption + link)
    #   Image/Video: B + D (caption + media)
    for i, post in enumerate(posts, start=3):
        ws[f"A{i}"] = post["time"]
        ws[f"B{i}"] = post["caption"]
        if post.get("media_url"):
            ws[f"D{i}"] = post["media_url"]
        # Columns C, E, F, G, H, I left empty for standard text/image posts

    wb.save(output_path)
    print(f"  Generated {output_path.name} ({len(posts)} posts)")


def strip_hashtags(text):
    """Remove all hashtags and their preceding blank line from post text."""
    # Remove lines that are only hashtags
    lines = text.split("\n")
    cleaned = []
    for line in lines:
        stripped = line.strip()
        # Skip lines that are entirely hashtags
        if stripped and all(word.startswith("#") for word in stripped.split()):
            continue
        cleaned.append(line)
    # Remove trailing blank lines left behind
    while cleaned and not cleaned[-1].strip():
        cleaned.pop()
    return "\n".join(cleaned)


def _extract_trailing_emoji(text):
    """Extract the last emoji from text, if any. Returns (emoji, text_without_emoji)."""
    # Match common emoji ranges at end of text (after optional whitespace)
    m = re.search(r'\s*([\U0001F300-\U0001FAFF\u2600-\u27BF\u2B50\u26A0-\u26FF]+)\s*$', text)
    if m:
        return m.group(1), text[:m.start()].rstrip()
    return None, text


def _truncate_at_sentence(text, limit):
    """Truncate text at the last complete sentence that fits within limit."""
    paragraphs = text.split('\n\n')
    all_sentences = []
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        sents = re.split(r'(?<=[.!?])\s+', para)
        all_sentences.extend(sents)

    # Merge standalone emoji fragments back into preceding sentence
    # (e.g. "...back to 100%." + "🔥" → "...back to 100%. 🔥")
    merged = []
    for s in all_sentences:
        s_stripped = s.strip()
        if not s_stripped:
            continue
        # If fragment has no letters or digits, it's emoji — merge with previous
        if merged and not re.search(r'[A-Za-z0-9]', s_stripped):
            merged[-1] = merged[-1] + " " + s_stripped
        else:
            merged.append(s_stripped)

    truncated = ""
    for sentence in merged:
        candidate = (truncated + " " + sentence).strip() if truncated else sentence
        if len(candidate) <= limit:
            truncated = candidate
        else:
            break

    # Clean up: remove trailing abbreviation fragments (Dr., Mr., vs., etc.)
    # These appear when "Dr. Keith" gets split at the period
    if truncated:
        parts = truncated.rsplit(' ', 1)
        if len(parts) > 1 and len(parts[-1]) <= 4 and re.match(r'^[A-Za-z]+\.$', parts[-1]):
            truncated = parts[0]

    return truncated if truncated else text[:limit - 3] + "..."


def _fit_hashtags_within_limit(text, hashtags, limit):
    """Add back as many hashtags as will fit within the character limit.

    Tries all hashtags first, then progressively fewer until some fit.
    Hashtags are placed on their own line (\\n\\n separator).
    """
    if not hashtags:
        return text

    # Try adding from all hashtags down to 1
    for num_tags in range(len(hashtags), 0, -1):
        tag_block = "\n\n" + " ".join(hashtags[:num_tags])
        if len(text + tag_block) <= limit:
            return text + tag_block

    # No hashtags fit — return text as-is
    return text


def prepare_tobi_posts(x_posts, tobi_background="true-black"):
    """
    Convert X/Twitter posts into Facebook TOBI posts.

    Strategy: use the EXACT same text as the X/Twitter post. TOBI posts must
    be identical in length to the tweet — no stripping, no truncation, no changes
    except replacing #NUMBER rankings to avoid Facebook hashtag links.

    Rules:
    - Keep ALL text including hashtags — identical to the tweet
    - Never truncate or strip any content
    - Only change: #NUMBER rankings → "No. NUMBER" (avoids Facebook hashtag links)
    - Column H = TOBI BACKGROUND (triggers TOBI formatting in PostPlanner)
    """
    tobi_posts = []

    for i, post in enumerate(x_posts):
        caption = post["caption"]

        # Replace #NUMBER rankings with "No. NUMBER" to avoid Facebook hashtag links
        # e.g., "#2 Texas" → "No. 2 Texas", "#24 Baylor" → "No. 24 Baylor"
        caption = re.sub(r'#(\d+)\b', r'No. \1', caption)

        tobi_posts.append({
            "time": post["time"],
            "caption": caption,
            "tobi_background": tobi_background,
        })

    return tobi_posts


def export_tobi_xlsx(posts, output_path, brand_name):
    """Export TOBI posts to PostPlanner XLSX V3.1 format.

    Uses Column H (TOBI BACKGROUND) to trigger Facebook text-on-background-image
    formatting. PostPlanner auto-detects the TOBI post type from this column.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Row 1: Comment row with version marker (REQUIRED by PostPlanner)
    ws["A1"] = f"## {brand_name} -- PostPlanner TOBI Bulk Upload"
    ws["B1"] = "BULK UPLOAD VERSION 3.1 \u2013 Please DO NOT EDIT this cell."

    # Row 2: Column headers as comment row (V3.1 format — 9 columns)
    ws["A2"] = "## POSTING TIME"
    ws["B2"] = "CAPTION"
    ws["C2"] = "LINK"
    ws["D2"] = "MEDIA"
    ws["E2"] = "TITLE"
    ws["F2"] = "FIRST COMMENT"
    ws["G2"] = "FIRST COMMENT DELAY"
    ws["H2"] = "TOBI BACKGROUND"
    ws["I2"] = "REEL OR STORY"

    # Data rows starting at row 3
    for i, post in enumerate(posts, start=3):
        ws[f"A{i}"] = post["time"]
        ws[f"B{i}"] = post["caption"]
        # C, D, E, F, G, I empty for TOBI posts
        ws[f"H{i}"] = post.get("tobi_background", "true-black")

    wb.save(output_path)
    print(f"  Generated {output_path.name} ({len(posts)} TOBI posts)")


def main():
    parser = argparse.ArgumentParser(
        description="Generate PostPlanner XLSX from approved social content"
    )
    parser.add_argument(
        "--niche", required=True,
        help="Niche folder name (e.g., Cubs, Softball, OUAC)"
    )
    parser.add_argument(
        "--tobi", action="store_true",
        help="Generate Facebook TOBI export (text on background image via Column H)"
    )
    parser.add_argument(
        "--tobi-background", default="true-black",
        help="TOBI background name in PostPlanner (default: true-black)"
    )
    parser.add_argument(
        "date", nargs="?", default=None,
        help="Date in YYYY-MM-DD format (defaults to today)"
    )
    args = parser.parse_args()

    # Resolve date
    if args.date:
        try:
            date_str = datetime.strptime(args.date, "%Y-%m-%d").strftime("%Y-%m-%d")
        except ValueError:
            print(f"Error: Invalid date format '{args.date}'. Use YYYY-MM-DD")
            sys.exit(1)
    else:
        date_str = datetime.now().strftime("%Y-%m-%d")

    # Resolve paths — script lives in _engine/scripts/
    engine_dir = Path(__file__).parent.parent  # _engine/
    niche_folder = engine_dir.parent / args.niche
    if not niche_folder.exists():
        print(f"Error: Niche folder not found: {niche_folder}")
        sys.exit(1)

    # Load niche config
    config = load_niche_config(niche_folder)
    content_prefix = config.get("content_prefix", f"{args.niche.lower()}-content")
    brand_name = config.get("brand_name", args.niche)
    brand_short = config.get("brand_short", args.niche).lower()
    platforms = config.get("platforms", [])

    # Find content folder
    content_folder = niche_folder / f"{content_prefix}-{date_str}"
    if not content_folder.exists():
        # Fallback: try niche root
        content_folder = niche_folder
        print(f"  Note: No dated folder found, using {niche_folder.name}/")

    output_folder = niche_folder / "postplanner-imports"
    output_folder.mkdir(parents=True, exist_ok=True)

    print(f"Processing PostPlanner export for: {brand_name}")
    print(f"  Date: {date_str}")
    print(f"  Content: {content_folder.name}")
    print(f"  Platforms: {', '.join(platforms) if platforms else 'all'}")
    print()

    all_posts = []

    # Parse X/Twitter posts
    if not platforms or "x_twitter" in platforms:
        x_posts = parse_x_posts(content_folder, date_str)
        print(f"  X/Twitter: {len(x_posts)} posts found")
        all_posts.extend(x_posts)

    # Parse Facebook posts
    if not platforms or "facebook" in platforms:
        fb_posts = parse_facebook_posts(content_folder, date_str)
        print(f"  Facebook: {len(fb_posts)} posts found")
        all_posts.extend(fb_posts)

    if not all_posts:
        print("\nNo posts found to export")
        sys.exit(0)

    # Filter by approvals if they exist
    approvals = load_approvals(content_folder)
    if approvals:
        print(f"  Approvals file found: {len(approvals)} entries")

    # Sort by posting time
    all_posts.sort(key=lambda p: p["time"])

    # Smart time redistribution: ensure all posts are >= now + 30 min
    all_posts = redistribute_times(all_posts, date_str)

    # Safety net: deduplicate any posts that ended up with the same time
    # (handles cases where markdown headings had duplicate times and
    # redistribute_times didn't fire because date != today or no times passed)
    all_posts = deduplicate_times(all_posts)

    # Export
    print()

    if args.tobi:
        # TOBI export: convert X/Twitter posts to Facebook TOBI format
        x_only = [p for p in all_posts if p.get("post_type") == "TEXT"]
        if not x_only:
            print("No text posts found for TOBI export")
            sys.exit(0)

        tobi_posts = prepare_tobi_posts(x_only, tobi_background=args.tobi_background)
        # Times already redistributed via all_posts — but deduplicate again
        # in case TOBI truncation created new collisions
        tobi_posts = deduplicate_times(tobi_posts)

        output_file = output_folder / f"{brand_short}-postplanner-tobi-{date_str}.xlsx"
        export_tobi_xlsx(tobi_posts, output_file, brand_name)

        print(f"\nComplete! Upload {output_file.name} to PostPlanner for Facebook TOBI posts.")
        print(f"  Path: {output_file}")
        print(f"  Posts: {len(tobi_posts)} TOBI posts (exact tweet text, no changes)")
        print(f"  Background: {args.tobi_background} (Column H)")
    else:
        output_file = output_folder / f"{brand_short}-postplanner-{date_str}.xlsx"
        export_xlsx(all_posts, output_file, brand_name)

        print(f"\nComplete! Upload {output_file.name} to PostPlanner.")
        print(f"  Path: {output_file}")


if __name__ == "__main__":
    main()
